from flask import Blueprint, abort, render_template, request, redirect, url_for, flash, jsonify, send_from_directory, session
from flask_login import login_required, current_user
from app import db
from app.models import suMenu, User, Permission, UserPermission
from app.forms import AccessForm, UserApprovalForm, UserPermissionForm, MenuForm
import os
import re
from utils.auth import get_allowed_permission_ids
import logging
from functools import wraps
from werkzeug.utils import secure_filename
from openpyxl import load_workbook

logging.basicConfig(level=logging.DEBUG)

admin_bp = Blueprint("admin", __name__, url_prefix="/admin")

# Custom decorator to require admin privileges
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # Check if the user is authenticated and is an admin
        if not current_user.is_admin():
            abort(403)  # Return 403 Forbidden error if not an admin
        return f(*args, **kwargs)
    return decorated_function

# Super user
@admin_bp.route("/user-approval", methods=["GET", "POST"])
@login_required
@admin_required
def user_approval():
    form = UserApprovalForm()
    
    if request.method == "GET":
        applicants = User.query.filter(
            User.approved.is_(None),  # Check if `approved` is None
            User.name.isnot(None),    # Ensure `name` is not None
            User.name != ''           # Ensure `name` is not an empty string
        ).all()
        filepath = os.path.join("uploads", "data-nik.xlsx")
        niks = []
        if os.path.exists(filepath):
            workbook = load_workbook(filepath)
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=1, max_col=1):
                cell_value = row[0].value
                if cell_value is not None:
                    niks.append(str(cell_value).replace('.0', '')[:10])
        return render_template("user_approval.html", applicants=applicants, form=form, niks=niks)
    
    elif request.method == "POST":
        if form.validate_on_submit():
            user = User.query.get(form.user_id.data)
            if form.approve.data:
                logging.info("Approved")
                user.approved = True
                flash(f"User {user.name} approved!", "success")
            elif form.reject.data:
                logging.info("Rejected")
                user.approved = False
                flash(f"User {user.name} rejected!", "danger")
            db.session.commit()
            return redirect(url_for("admin.user_approval"))
        else:
            flash("Form validation failed!", "danger")
        return redirect(url_for("admin.user_approval"))

@admin_bp.route("/user-permission", methods=["GET", "POST"])
@admin_bp.route("/user-permission/<user_id>", methods=["GET"])
@login_required
@admin_required
def user_permission(user_id=None):
    form = UserPermissionForm()
    if user_id:
        user_permissions = UserPermission.query.filter_by(user_id=user_id).all()
        permission_ids = [up.permission_id for up in user_permissions]
        return jsonify({"permissions": permission_ids})
    
    if request.method == "POST":
        if form.validate_on_submit():
            user_id = form.user_id.data
            selected_permissions = form.permission_id.data
            user = User.query.get(user_id)
            user_permissions = UserPermission.query.join(User).filter(User.id==UserPermission.user_id, UserPermission.user_id==user_id).all()
            for permission_id in selected_permissions:
                if permission_id not in [up.permission_id for up in user_permissions]:
                    new_permission = UserPermission(user_id=user_id, permission_id=permission_id)
                    db.session.add(new_permission)
            for user_permission in user_permissions:
                if user_permission.permission_id not in selected_permissions:
                    db.session.delete(user_permission)
            db.session.commit()
            flash(f"Permissions updated for {user.name}", "success")
        else:
            flash("Form validation failed", "danger")

        return redirect(url_for("admin.user_permission"))

    # For GET requests
    users = User.query.filter(User.approved == True).all()
    users.sort(key=lambda x: x.name)
    return render_template("user_permission.html", users=users, form=form)

@admin_bp.route("/", methods=["GET"])
@admin_bp.route("/menu", methods=["GET", "POST"])
@login_required
@admin_required
def menu():
    form = MenuForm()
    if request.method == "POST":
        print("Permission ID: ", form.permission_id.data)
        if form.permission_id.data == 0:
            if form.permission_name.data is None or form.permission_name.data == "":
                flash("Please select a permission!", "danger")
                return redirect(url_for("admin.menu"))            
            else:
                permission = Permission(name=form.permission_name.data, description=form.permission_description.data)
                db.session.add(permission)
                db.session.commit()
                permission_id = permission.id
        else:
            permission_id = form.permission_id.data

        print("Permission ID: ", permission_id)

        if form.is_submitted():                        
            title = form.title.data
            url = form.url.data or None
            file = form.file.data
            new_menu = suMenu(title=title, url=url, permission_id=permission_id, file=file.read())
            db.session.add(new_menu)
            db.session.commit()
            flash("Menu item added successfully!", "success")
        else:
            flash("Form validation failed!", "danger")
        return redirect(url_for("admin.menu"))
    permission_ids = get_allowed_permission_ids()
    menus = suMenu.query.filter(suMenu.permission_id.in_(permission_ids)).order_by(suMenu.title.asc()).all()
    form.permission_id.choices.insert(0, (0, "Select Permission"))
    form.permission_name.data = None
    return render_template("menu.html", form=form, menus=menus)

@admin_bp.route('/set_session', methods=['POST'])
def set_session():
    permission_id = request.form.get('id')    
    if permission_id:
        permission = Permission.query.get(permission_id)
        session['permission_name'] = permission.name
        session['permission_id'] = permission_id
        return jsonify({'status': 'success'}), 200
    return jsonify({'status': 'error', 'message': 'No ID provided'}), 400

@admin_bp.app_context_processor
def inject_session_permission():
    return dict(session_permission=session.get('permission_id', ''), permission_name=session.get('permission_name', ''))

@admin_bp.route('/upload-niks', methods=['GET', 'POST'])
def upload_niks():
    if request.method == "POST":
        file = request.files['file']
        if file:
            if file.filename.split('.')[-1] not in ['xlsx', 'xls']:
                flash("Invalid file type. Please upload an XLS or XLSX file.", "danger")
                return redirect(url_for("admin.upload_niks"))

            filename = "data-nik.xlsx"
            file_path = os.path.join('uploads', filename)
            file.save(file_path)            
            try:
                # Sync User NIK and NIK from Excel
                users = User.query.all()
                niks = [user.id for user in users]
                workbook = load_workbook(file_path)
                sheet = workbook.active
                niks_from_excel = set() 
                for row in sheet.iter_rows(min_row=1, max_col=1):
                    cell_value = row[0].value
                    if cell_value is not None:
                        nik_str = str(cell_value).replace('.0', '')[:10]
                        niks_from_excel.add(nik_str)
                niks_to_remove = set(niks) - niks_from_excel
                niks_to_add = niks_from_excel - set(niks)
                for nik in niks_to_remove:
                    user_to_remove = User.query.filter_by(id=nik).first()
                    if user_to_remove:
                        db.session.delete(user_to_remove)

                for nik in niks_to_add:
                    new_user = User(id=nik)
                    db.session.add(new_user)
                db.session.commit()

                # Auto Reject User
                applicants = User.query.filter(
                    User.approved.is_(None), 
                    User.name.isnot(None),
                    User.name != ''
                ).all()

                for applicant in applicants:
                    if applicant.id not in niks_from_excel:
                        applicant.approved = False
                db.session.commit()                

                flash("NIKs processed successfully. Updated database accordingly.", "success")
            except Exception as e:
                db.session.rollback()
                flash(f"An error occurred while processing the file: {str(e)}", "danger")
            return redirect(url_for("admin.upload_niks"))
        else:
            flash("No file selected", "danger")
            return redirect(url_for("admin.upload_niks"))
    return redirect(url_for("admin.user_approval"))